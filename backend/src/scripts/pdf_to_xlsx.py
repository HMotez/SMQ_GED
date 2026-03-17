"""
pdf_to_xlsx.py — Extract tables and text from PDF to Excel using pdfplumber + openpyxl
Usage: python3 pdf_to_xlsx.py <input.pdf> <output.xlsx>
"""
import sys
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

if len(sys.argv) != 3:
    print("Usage: pdf_to_xlsx.py <input.pdf> <output.xlsx>", file=sys.stderr)
    sys.exit(1)

src  = sys.argv[1]
dest = sys.argv[2]

try:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    with pdfplumber.open(src) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            ws = wb.create_sheet(title=f"Page {page_num}")

            # Try to extract tables first
            tables = page.extract_tables()
            if tables:
                row_idx = 1
                for table in tables:
                    for row in table:
                        for col_idx, cell in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=cell or "")
                        row_idx += 1
                    row_idx += 1  # blank row between tables
            else:
                # Fall back to plain text extraction (one line per row)
                text = page.extract_text() or ""
                for row_idx, line in enumerate(text.splitlines(), start=1):
                    ws.cell(row=row_idx, column=1, value=line)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Page 1")
        ws.cell(row=1, column=1, value="Aucun contenu extractible détecté.")

    wb.save(dest)
    print(f"OK: {dest}")

except Exception as e:
    print(f"ERROR: {e}", file=sys.stderr)
    sys.exit(1)
